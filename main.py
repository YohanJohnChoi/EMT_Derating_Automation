"""
BOM Parser (Excel -> Template Derating) - Final All-in-One

✅ 포함된 요구사항
1) 템플릿 시트에서 레코드(파트) 수가 늘어 추가 행/블록 생성 시,
   템플릿 레코드 블록(1개 파트에 해당하는 여러 행 구조)의 "셀 병합(merge)" 구조까지 동일하게 복제
   -> 파트명/값이 병합 셀에 깨져 들어가는 문제 방지

2) EXE(또는 스크립트)와 같은 폴더에서 파일 자동 감지하여 GUI 입력칸 프리필
   - 파일명에 "TEMPLATE" 포함된 .xlsx -> Template 기본값
   - 파일명에 "LOOKUPTABLE" 포함된 .xlsx -> LookupTable 기본값
   - 유저가 Browse로 언제든 변경 가능

3) GUI는 한 화면에서 BOM/Template/Lookup/Output을 언제든 변경 가능 (Run 버튼 실행)

4) BOM 열 순서가 바뀌어도 헤더명 기반 탐색으로 정상 동작 (헤더명이 유지되는 경우)

5) 리포트(txt): 중복 Ref, 정격값 누락(부분 포함), 라우팅 히트 기록
   - Connector는 V_RATED/V_MAX 누락은 리포트에 기록하지 않음 (전류 I만 체크)

6) ✅ 특정 시트에서 해당되는 파트가 0개인 경우,
   템플릿에 적힌 예시(첫 레코드) 부품 내역이 남지 않도록 "값"을 공란 처리(서식/병합/수식 구조 유지)

빌드:
pyinstaller --onefile --windowed --icon app.ico bom_parser_app.py
"""

import re
import sys
import argparse
from dataclasses import dataclass
from pathlib import Path
from collections import defaultdict
from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator

import tkinter as tk
from tkinter import filedialog, messagebox


# =========================
# 기본 유틸
# =========================
def normalize_text(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\u00A0", " ").replace("\u200B", "").replace("\ufeff", "")
    return s.strip()


def normalize_category(v) -> str:
    s = normalize_text(v).upper()
    s = re.sub(r"\s+", "", s)
    return s


def normalize_part(v) -> str:
    return normalize_text(v)


def normalize_subcategory(v) -> str:
    s = normalize_text(v)
    if not s:
        return ""
    u = s.upper()
    if u in ("(BLANK)", "BLANK", "NONE", "N/A"):
        return ""
    return s


def normalize_ref_list(location_value):
    s = normalize_text(location_value)
    if not s:
        return []
    # 중복 Ref는 유지
    return [p.strip() for p in s.split(",") if p.strip()]


def ref_sort_key(ref):
    s = normalize_text(ref).upper()
    m = re.match(r"^([A-Z]+)\s*0*([0-9]+)(.*)$", s)
    if m:
        return (m.group(1), int(m.group(2)), m.group(3).strip())
    return (s, 10**12, "")


def extract_voltage(detail_spec):
    if not detail_spec:
        return ""
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*V\b", str(detail_spec), flags=re.IGNORECASE)
    return f"{matches[-1]}V" if matches else ""


def safe_set(ws, r, c, value):
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


# =========================
# 실행 폴더(Exe/Script) 기반 자동 감지
# =========================
def get_app_dir() -> Path:
    """
    exe 실행 시: exe가 있는 폴더
    py 실행 시: 스크립트가 있는 폴더
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def autodetect_default_files(app_dir: Path):
    """
    app_dir에서 TEMPLATE/LOOKUPTABLE 포함된 xlsx 자동 탐지
    - 여러 개면 "수정시간 최신" 우선
    """
    template = None
    lookup = None

    xlsx = sorted(app_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for p in xlsx:
        name_u = p.name.upper()
        if template is None and "TEMPLATE" in name_u:
            template = p
        if lookup is None and "LOOKUPTABLE" in name_u:
            lookup = p
        if template and lookup:
            break
    return template, lookup


# =========================
# 템플릿 레이아웃 감지/복사/클리어
# =========================
def find_data_start_and_step(ws, scan_rows=300):
    """
    Column A에서 값이 1인 최초 행을 레코드 시작으로 보고,
    해당 행이 포함된 merged 범위를 기준으로 레코드 step(행수)를 추정.
    """
    start_row = None
    for r in range(1, scan_rows + 1):
        if ws.cell(r, 1).value == 1:
            start_row = r
            break
    if start_row is None:
        return 6, 1

    step = 1
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= 1 <= rng.max_col and rng.min_row <= start_row <= rng.max_row:
            if rng.min_col == 1:
                step = max(step, rng.max_row - rng.min_row + 1)
    return start_row, step


def copy_row_with_formula_translate(ws, src_row, dst_row, max_col):
    """
    셀 스타일/값/수식을 복사하면서 수식은 행 이동에 맞게 translate.
    (병합은 별도 함수로 처리)
    """
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if isinstance(dst, MergedCell):
            continue

        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)

        v = src.value
        if isinstance(v, str) and v.startswith("="):
            dst.value = Translator(v, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = v


def clear_records(ws, start_row, step, cols, n_records=1200):
    """
    레코드 데이터 영역 값만 clear (template 첫 레코드는 유지)
    """
    max_r = ws.max_row
    for i in range(n_records):
        r0 = start_row + i * step
        if r0 > max_r:
            break
        for rr in range(r0, min(r0 + step, max_r + 1)):
            for c in cols:
                cell = ws.cell(rr, c)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None


def clear_first_record_values(ws, start_row, step, cfg):
    """
    ✅ 특정 시트에 실제 출력할 부품이 0개인 경우,
    템플릿의 첫 레코드(예시 데이터) 값을 공란으로 만들기(서식/병합/수식은 유지).
    """
    cols = [1, 2, 3, cfg["spec_col"], cfg["actual_col"]]
    if cfg.get("detail_col"):
        cols.append(cfg["detail_col"])

    for rr in range(start_row, start_row + step):
        for c in cols:
            cell = ws.cell(rr, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


# =========================
# ✅ 병합(merge) 복제 지원
# =========================
def get_record_merges(ws, start_row, step):
    """
    템플릿 레코드(1개 블록) 영역에 포함되는 merge 범위를 수집.
    반환: [(min_row, min_col, max_row, max_col), ...]
    """
    r0 = start_row
    r1 = start_row + step - 1
    merges = []
    for rng in ws.merged_cells.ranges:
        if rng.max_row < r0 or rng.min_row > r1:
            continue
        # 레코드 블록 밖으로 걸치는 merge는 블록 복사에서 위험 -> 제외
        if rng.min_row < r0 or rng.max_row > r1:
            continue
        merges.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
    return merges


def unmerge_block(ws, block_start_row, step, max_col=None):
    """
    새 레코드 블록 영역에 걸쳐있는 merge를 모두 해제
    """
    r0 = block_start_row
    r1 = block_start_row + step - 1
    ranges = list(ws.merged_cells.ranges)
    for rng in ranges:
        if rng.max_row < r0 or rng.min_row > r1:
            continue
        if max_col is not None and (rng.min_col > max_col or rng.max_col < 1):
            continue
        ws.unmerge_cells(str(rng))


def apply_record_merges(ws, template_merges, row_offset):
    """
    템플릿 레코드 merge들을 row_offset만큼 이동시켜 동일하게 merge 적용
    """
    for (min_r, min_c, max_r, max_c) in template_merges:
        ws.merge_cells(
            start_row=min_r + row_offset,
            start_column=min_c,
            end_row=max_r + row_offset,
            end_column=max_c
        )


# =========================
# 룩업(Field) 정규화
# =========================
def normalize_field(raw_field: str) -> str:
    s = normalize_text(raw_field).upper()
    if not s:
        return ""
    s = re.sub(r"[^A-Z0-9_]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    key = s.replace("_", "")

    synonyms = {
        "IRATED": "I_RATED",
        "VRATED": "V_RATED",
        "IMAX": "I_MAX",
        "VMAX": "V_MAX",
        "PMAX": "P_MAX",
        "PRATED": "P_MAX",
        "POWERMAX": "P_MAX",
        "POWERRATED": "P_MAX",

        "VRWM": "VRWM",
        "VRRM": "V_MAX",
        "VBRVPT": "VBR_VPT",
        "VBR_VPT": "VBR_VPT",
        "VBR": "VBR_VPT",

        "VDDMAX": "V_MAX",
        "VINMAX": "V_MAX",
    }
    return synonyms.get(key, s)


@dataclass
class RatingRec:
    field: str
    value: str
    unit: str
    priority: int = 1


def format_value_unit(value: str, unit: str) -> str:
    v = normalize_text(value)
    u = normalize_text(unit)
    if not v:
        return ""
    return f"{v}{u}"


def load_resistor_prefix_rules(lk_wb):
    if RES_PREFIX_SHEET not in lk_wb.sheetnames:
        return {}

    ws = lk_wb[RES_PREFIX_SHEET]
    hdr = {normalize_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    for req in ["Prefix", "Rating_Value", "Rating_Unit"]:
        if req not in hdr:
            raise ValueError(f"{RES_PREFIX_SHEET} 시트에 필요한 헤더가 없습니다: {req}")

    rules = defaultdict(list)
    has_vendor = "Vendor" in hdr
    has_priority = "Priority" in hdr

    for r in range(2, ws.max_row + 1):
        prefix = normalize_text(ws.cell(r, hdr["Prefix"]).value).upper()
        if not prefix:
            continue

        raw_val = ws.cell(r, hdr["Rating_Value"]).value
        raw_unit = ws.cell(r, hdr["Rating_Unit"]).value
        if raw_val is None and raw_unit is None:
            continue

        vendor = ""
        if has_vendor:
            vendor = normalize_text(ws.cell(r, hdr["Vendor"]).value).upper()

        pr = 1
        if has_priority:
            pv = ws.cell(r, hdr["Priority"]).value
            if isinstance(pv, (int, float)):
                pr = int(pv)

        rules[prefix].append({
            "vendor": vendor,
            "priority": pr,
            "value_unit": format_value_unit(raw_val, raw_unit),
        })

    return rules


def pick_resistor_prefix_rating(part_name: str, prefix_rules: dict) -> str:
    if not prefix_rules:
        return ""
    s = normalize_text(part_name)
    if not s:
        return ""
    prefix = s[:5].upper()
    candidates = prefix_rules.get(prefix, [])
    if not candidates:
        return ""

    walsin = [c for c in candidates if c["vendor"] == "WALSIN"]
    pool = walsin if walsin else candidates
    pool_sorted = sorted(pool, key=lambda x: (x["priority"], x["value_unit"]))
    return pool_sorted[0]["value_unit"] if pool_sorted else ""


# =========================
# 시트/규칙
# =========================
SHEET_CFG = {
    "Resistor": {"detail_col": 4, "spec_col": 5, "actual_col": 6},
    "Capacitor": {"detail_col": 4, "spec_col": 5, "actual_col": 6},
    "Inductor": {"detail_col": 4, "spec_col": 5, "actual_col": 6},
    "Diode(ESD_Zener_Surge)": {"detail_col": 4, "spec_col": 5, "actual_col": 6},
    "Diode(Schottky_switching)": {"detail_col": 4, "spec_col": 5, "actual_col": 6},
    "FET&TR": {"detail_col": 4, "spec_col": 5, "actual_col": 6},
    "DCDC & LDO": {"detail_col": 4, "spec_col": 5, "actual_col": 6},
    "IC": {"detail_col": None, "spec_col": 4, "actual_col": 5},
    "Connector": {"detail_col": None, "spec_col": 4, "actual_col": 5},
}
MANAGED_SHEETS = list(SHEET_CFG.keys())
UNCLASS_SHEET = "미분류"
RES_PREFIX_SHEET = "RESISTOR_PREFIX"

RATING_SLOTS_BY_SHEET = {
    "Diode(Schottky_switching)": ["V_MAX", "I_MAX"],
    "Diode(ESD_Zener_Surge)": ["VRWM", "VBR_VPT"],
    "FET&TR": ["V_MAX", "I_MAX"],
    "DCDC & LDO": ["V_MAX", "I_MAX", "P_MAX"],
}

FIELD_ORDER_BY_CATEGORY = {
    "RESISTOR": ["P_MAX", "POWER_MAX", "W_MAX"],
    "INDUCTOR": ["I_RATED", "I_MAX", "CURRENT", "DCR"],
    "IC": ["V_MAX", "V_RATED", "I_MAX", "P_MAX"],
    "DIODE": ["VRWM", "VBR_VPT", "V_MAX", "I_MAX"],
    "TR": ["V_MAX", "I_MAX"],
    "CONNECTOR": ["I_RATED", "I_MAX"],  # Connector V는 필요없음
    "CAPACITOR": ["V_RATED", "V_MAX"],
}

SUGGEST_ALT_FIELDS = {
    "V_MAX": ["V_RATED", "VRRM", "VRWM"],
    "I_MAX": ["I_RATED"],
    "P_MAX": ["POWER_MAX", "PRATED"],
    "VRWM": ["V_MAX", "V_RATED"],
    "VBR_VPT": ["VBR"],
}


def suggest_alternatives(missing_field: str, available_fields: set) -> list:
    return [cand for cand in SUGGEST_ALT_FIELDS.get(missing_field, []) if cand in available_fields]


def build_connector_spec(rating_map_for_part: dict) -> str:
    i = None
    for f in ["I_RATED", "I_MAX"]:
        if f in rating_map_for_part and rating_map_for_part[f]:
            i = rating_map_for_part[f]
            break
    return i or ""


# =========================
# 리포트
# =========================
def write_issue_report(report_path: Path, duplicate_refs: dict, rating_issues: list, routed_items: list):
    lines = []
    lines.append("=== BOM Parsing Issues Report ===")
    lines.append(f"- Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- Report file: {report_path.name}")
    lines.append("")

    lines.append("[0] Routing Hits (기본 시트가 아닌 다른 시트로 라우팅된 항목)")
    if not routed_items:
        lines.append("  - None")
    else:
        for it in sorted(routed_items, key=lambda x: ref_sort_key(x["ref"])):
            lines.append(
                f"  - Ref={it['ref']}, Part={it['part']}, Category={it['cat']}, Subcategory={it['sub'] or '(blank)'}"
            )
            lines.append(
                f"      BOM_row={it['bom_row']}, BaseSheet={it['base_sheet']} -> TargetSheet={it['sheet']}"
            )
    lines.append("")

    lines.append("[1] Duplicate References (중복 Ref)")
    if not duplicate_refs:
        lines.append("  - None")
    else:
        for ref in sorted(duplicate_refs.keys(), key=ref_sort_key):
            occs = duplicate_refs[ref]
            lines.append(f"  - {ref} (count={len(occs)})")
            for o in occs:
                lines.append(
                    f"      * BOM_row={o['bom_row']}, Category={o['cat']}, Subcategory={o['sub'] or '(blank)'}"
                    f", Part={o['part']}, TargetSheet={o['sheet']}"
                )
    lines.append("")

    lines.append("[2] Missing Ratings (정격값 공란/부분 공란) + Suggestions")
    if not rating_issues:
        lines.append("  - None")
    else:
        for it in rating_issues:
            miss = ", ".join(it["missing_fields"]) if it["missing_fields"] else "(unknown)"
            lines.append(
                f"  - Sheet={it['sheet']}, Ref={it['ref']}, Part={it['part']}, "
                f"Category={it['cat']}, Subcategory={it['sub'] or '(blank)'}, BOM_row={it['bom_row']}"
            )
            lines.append(f"      MissingFields: {miss}")
            lines.append(f"      LookupHasAnyRatingForPart: {it['lookup_has_any']}")
            if it["lookup_has_any"]:
                lines.append(f"      AvailableCanonicalFields: {', '.join(sorted(it['available_fields'])) or '(none)'}")
                if it.get("available_raw_fields"):
                    lines.append(f"      AvailableRawFields: {', '.join(sorted(it['available_raw_fields']))}")
                for mf in it["missing_fields"]:
                    alts = it.get("suggestions", {}).get(mf, [])
                    if alts:
                        lines.append(f"      SuggestFor[{mf}]: use {', '.join(alts)} (if acceptable)")
            lines.append("")
    report_path.write_text("\n".join(lines), encoding="utf-8")


def write_unclassified_sheet(tpl_wb, ws_bom, unclassified_rows, sheet_name: str):
    if sheet_name in tpl_wb.sheetnames:
        ws = tpl_wb[sheet_name]
    else:
        ws = tpl_wb.create_sheet(sheet_name)

    max_col = ws_bom.max_column
    for c in range(1, max_col + 1):
        ws.cell(1, c).value = ws_bom.cell(1, c).value

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for i, row in enumerate(sorted(unclassified_rows, key=lambda x: x["bom_row"]), start=2):
        vals = row["values"]
        for c, val in enumerate(vals, start=1):
            ws.cell(i, c).value = val


# =========================
# 파서 실행
# =========================
def run_parser(bom_path: Path, template_path: Path, lookup_path: Path, out_xlsx: Path, out_txt: Path):
    # --- lookup ---
    lk_wb = openpyxl.load_workbook(lookup_path, data_only=True)
    if "TABLE" not in lk_wb.sheetnames or "ROUTING_RULES" not in lk_wb.sheetnames:
        raise ValueError("룩업테이블에는 'TABLE'과 'ROUTING_RULES' 시트가 필요합니다.")

    ws_table = lk_wb["TABLE"]
    ws_rules = lk_wb["ROUTING_RULES"]

    table_hdr = {normalize_text(ws_table.cell(1, c).value): c for c in range(1, ws_table.max_column + 1)}
    rules_hdr = {normalize_text(ws_rules.cell(1, c).value): c for c in range(1, ws_rules.max_column + 1)}

    for req in ["Category", "Subcategory", "Part_Name", "Rating_Field", "Rating_Value", "Rating_Unit"]:
        if req not in table_hdr:
            raise ValueError(f"LOOKUP TABLE에 필요한 헤더가 없습니다: {req}")
    for req in ["Category", "Subcategory", "Output_Sheet"]:
        if req not in rules_hdr:
            raise ValueError(f"ROUTING_RULES에 필요한 헤더가 없습니다: {req}")

    ratings = defaultdict(list)
    subcat_map = {}
    raw_field_map = defaultdict(set)
    part_to_cats = defaultdict(set)
    has_priority = "Priority" in table_hdr

    for r in range(2, ws_table.max_row + 1):
        cat = normalize_category(ws_table.cell(r, table_hdr["Category"]).value)
        sub = normalize_subcategory(ws_table.cell(r, table_hdr["Subcategory"]).value)
        part = normalize_part(ws_table.cell(r, table_hdr["Part_Name"]).value)

        raw_field = ws_table.cell(r, table_hdr["Rating_Field"]).value
        field = normalize_field(raw_field)

        raw_val = ws_table.cell(r, table_hdr["Rating_Value"]).value
        raw_unit = ws_table.cell(r, table_hdr["Rating_Unit"]).value

        pr = 1
        if has_priority:
            pv = ws_table.cell(r, table_hdr["Priority"]).value
            if isinstance(pv, (int, float)):
                pr = int(pv)

        if not cat or not part:
            continue
        part_to_cats[part].add(cat)

        key = (cat, part)
        if key not in subcat_map or (not subcat_map[key] and sub):
            subcat_map[key] = sub
        if raw_field is not None:
            raw_field_map[key].add(normalize_text(raw_field))

        if field:
            val = "" if raw_val is None else str(raw_val).strip()
            unit = normalize_text(raw_unit)
            ratings[key].append(RatingRec(field=field, value=val, unit=unit, priority=pr))

    routing = {}
    for r in range(2, ws_rules.max_row + 1):
        cat = normalize_category(ws_rules.cell(r, rules_hdr["Category"]).value)
        sub = normalize_subcategory(ws_rules.cell(r, rules_hdr["Subcategory"]).value)
        out_sheet = normalize_text(ws_rules.cell(r, rules_hdr["Output_Sheet"]).value)
        if cat and out_sheet:
            routing[(cat, sub)] = out_sheet

    rating_map = defaultdict(dict)
    for key, recs in ratings.items():
        recs_sorted = sorted(recs, key=lambda x: (x.priority, x.field))
        for rec in recs_sorted:
            if rec.field in rating_map[key]:
                continue
            rating_map[key][rec.field] = format_value_unit(rec.value, rec.unit)

    resistor_prefix_rules = load_resistor_prefix_rules(lk_wb)

    # --- template ---
    tpl_wb = openpyxl.load_workbook(template_path)
    for s in MANAGED_SHEETS:
        if s not in tpl_wb.sheetnames:
            raise ValueError(f"템플릿에 시트가 없습니다: '{s}'")

    layouts = {}
    record_merges = {}
    for s in MANAGED_SHEETS:
        ws = tpl_wb[s]
        start_row, step = find_data_start_and_step(ws)
        layouts[s] = (start_row, step)
        record_merges[s] = get_record_merges(ws, start_row, step)

    # --- BOM ---
    wb_bom = openpyxl.load_workbook(bom_path, data_only=True)
    ws_bom = wb_bom.active

    bom_hdr = {normalize_text(ws_bom.cell(1, c).value): c for c in range(1, ws_bom.max_column + 1)}
    for req in ["품목명", "분류체계", "세부규격", "Location"]:
        if req not in bom_hdr:
            raise ValueError(f"BOM에 필요한 헤더가 없습니다: {req}")

    col_part = bom_hdr["품목명"]
    col_cls = bom_hdr["분류체계"]
    col_detail = bom_hdr["세부규격"]
    col_loc = bom_hdr["Location"]

    base_cat_to_sheet = {
        "RESISTOR": "Resistor",
        "CAPACITOR": "Capacitor",
        "INDUCTOR": "Inductor",
        "TR": "FET&TR",
        "CONNECTOR": "Connector",
        "DIODE": routing.get(("DIODE", ""), "Diode(ESD_Zener_Surge)"),
        "IC": routing.get(("IC", ""), "IC"),
    }

    out_groups = defaultdict(list)
    ignored = defaultdict(int)
    ref_occurrences = defaultdict(list)
    routed_items = []
    unclassified_rows = []

    for r in range(2, ws_bom.max_row + 1):
        raw_cat = ws_bom.cell(r, col_cls).value
        raw_part = ws_bom.cell(r, col_part).value
        raw_detail = ws_bom.cell(r, col_detail).value
        raw_loc = ws_bom.cell(r, col_loc).value

        if raw_cat is None and raw_part is None and raw_detail is None and raw_loc is None:
            continue

        cat = normalize_category(raw_cat)
        part = normalize_part(raw_part)
        detail = "" if raw_detail is None else normalize_text(raw_detail)

        if cat == "FILTER":
            candidates = []
            for c in part_to_cats.get(part, set()):
                sub_c = subcat_map.get((c, part), "")
                sheet_c = routing.get((c, sub_c), routing.get((c, ""), base_cat_to_sheet.get(c)))
                if sheet_c in MANAGED_SHEETS:
                    candidates.append(c)
            if len(candidates) == 1:
                cat = candidates[0]

        row_values = [ws_bom.cell(r, c).value for c in range(1, ws_bom.max_column + 1)]

        if not cat or cat not in base_cat_to_sheet:
            unclassified_rows.append({"bom_row": r, "values": row_values})
            continue

        base_sheet = base_cat_to_sheet[cat]
        sub = subcat_map.get((cat, part), "")
        sheet = routing.get((cat, sub), routing.get((cat, ""), base_sheet))

        if sheet not in MANAGED_SHEETS:
            unclassified_rows.append({"bom_row": r, "values": row_values})
            continue

        for ref in normalize_ref_list(raw_loc):
            item = {
                "bom_row": r,
                "cat": cat,
                "sub": sub,
                "ref": ref,
                "part": part,
                "detail": detail,
                "sheet": sheet,
                "base_sheet": base_sheet,
            }
            out_groups[sheet].append(item)

            ref_occurrences[ref].append({
                "bom_row": r,
                "cat": cat,
                "sub": sub,
                "part": part,
                "sheet": sheet,
            })

            if sheet != base_sheet:
                routed_items.append({
                    "bom_row": r,
                    "cat": cat,
                    "sub": sub,
                    "ref": ref,
                    "part": part,
                    "sheet": sheet,
                    "base_sheet": base_sheet,
                })

    duplicate_refs = {ref: occs for ref, occs in ref_occurrences.items() if len(occs) > 1}

    # --- clear template data area (keep first record) ---
    for sheet_name in MANAGED_SHEETS:
        ws = tpl_wb[sheet_name]
        start_row, step = layouts[sheet_name]
        cfg = SHEET_CFG[sheet_name]

        cols_to_clear = {1, 2, 3, cfg["spec_col"], cfg["actual_col"]}
        if cfg.get("detail_col"):
            cols_to_clear.add(cfg["detail_col"])

        clear_from = start_row + step
        clear_records(ws, clear_from, step, cols=sorted(cols_to_clear), n_records=1200)

    # --- write + rating issues ---
    rating_issues = []

    for sheet_name, items in out_groups.items():
        ws = tpl_wb[sheet_name]
        start_row, step = layouts[sheet_name]
        cfg = SHEET_CFG[sheet_name]
        max_col = ws.max_column

        items.sort(key=lambda x: ref_sort_key(x["ref"]))
        template_row = start_row

        for i, item in enumerate(items, start=1):
            record_start = start_row + (i - 1) * step

            # ✅ 병합 복제: 해제 -> 복사 -> 병합 적용
            unmerge_block(ws, record_start, step, max_col=max_col)

            for off in range(step):
                copy_row_with_formula_translate(ws, template_row + off, record_start + off, max_col=max_col)

            row_offset = record_start - template_row
            apply_record_merges(ws, record_merges[sheet_name], row_offset)

            # 기본 값 기입
            safe_set(ws, record_start, 1, i)
            safe_set(ws, record_start, 2, item["ref"])
            safe_set(ws, record_start, 3, item["part"])
            if cfg.get("detail_col"):
                safe_set(ws, record_start, cfg["detail_col"], item["detail"])
            safe_set(ws, record_start, cfg["actual_col"], None)

            # 정격값 채우기
            cat = item["cat"]
            part = item["part"]
            part_ratings = rating_map.get((cat, part), {})
            lookup_has_any = bool(part_ratings)
            available_fields = set(part_ratings.keys())
            available_raw_fields = raw_field_map.get((cat, part), set())

            # Capacitor: 세부규격에서 전압 추출(우선), 없으면 룩업
            if sheet_name == "Capacitor":
                cap_v = extract_voltage(item["detail"])
                if cap_v:
                    safe_set(ws, record_start, cfg["spec_col"], cap_v)
                else:
                    spec_one = ""
                    for f in FIELD_ORDER_BY_CATEGORY.get("CAPACITOR", []):
                        if f in part_ratings and part_ratings[f]:
                            spec_one = part_ratings[f]
                            break
                    safe_set(ws, record_start, cfg["spec_col"], spec_one if spec_one else None)
                    if not spec_one:
                        rating_issues.append({
                            "sheet": sheet_name,
                            "ref": item["ref"],
                            "part": part,
                            "cat": cat,
                            "sub": item["sub"],
                            "bom_row": item["bom_row"],
                            "missing_fields": ["(CAP_VOLTAGE)"],
                            "lookup_has_any": lookup_has_any,
                            "available_fields": sorted(available_fields),
                            "available_raw_fields": sorted(available_raw_fields),
                            "suggestions": {}
                        })
                continue

            # Connector: 전류만 표시/체크 (전압은 필요 없음)
            if sheet_name == "Connector":
                spec = build_connector_spec(part_ratings)
                safe_set(ws, record_start, cfg["spec_col"], spec if spec else None)

                i_present = any(part_ratings.get(f) for f in ["I_RATED", "I_MAX"])
                if not i_present:
                    rating_issues.append({
                        "sheet": sheet_name,
                        "ref": item["ref"],
                        "part": part,
                        "cat": cat,
                        "sub": item["sub"],
                        "bom_row": item["bom_row"],
                        "missing_fields": ["I_RATED/I_MAX"],
                        "lookup_has_any": lookup_has_any,
                        "available_fields": sorted(available_fields),
                        "available_raw_fields": sorted(available_raw_fields),
                        "suggestions": {}
                    })
                continue

            # 슬롯 규칙(여러 행에 정격값 순서대로 기입)
            if sheet_name == "Resistor":
                spec_one = pick_resistor_prefix_rating(part, resistor_prefix_rules)
                if not spec_one:
                    for f in FIELD_ORDER_BY_CATEGORY.get(cat, []):
                        if f in part_ratings and part_ratings[f]:
                            spec_one = part_ratings[f]
                            break
                safe_set(ws, record_start, cfg["spec_col"], spec_one if spec_one else None)

                if not spec_one:
                    suggestions = {}
                    if lookup_has_any:
                        for target in ["V_MAX", "I_MAX", "P_MAX", "V_RATED", "I_RATED"]:
                            alts = suggest_alternatives(target, available_fields)
                            if alts:
                                suggestions[target] = alts

                    rating_issues.append({
                        "sheet": sheet_name,
                        "ref": item["ref"],
                        "part": part,
                        "cat": cat,
                        "sub": item["sub"],
                        "bom_row": item["bom_row"],
                        "missing_fields": ["(NO_MATCHED_FIELD)"],
                        "lookup_has_any": lookup_has_any,
                        "available_fields": sorted(available_fields),
                        "available_raw_fields": sorted(available_raw_fields),
                        "suggestions": suggestions
                    })
                continue

            if sheet_name in RATING_SLOTS_BY_SHEET:
                slots = RATING_SLOTS_BY_SHEET[sheet_name]
                n = min(len(slots), step)

                missing = []
                suggestions = {}
                for idx in range(n):
                    field = slots[idx]
                    val = part_ratings.get(field, "")
                    safe_set(ws, record_start + idx, cfg["spec_col"], val if val else None)
                    if not val:
                        missing.append(field)
                        alts = suggest_alternatives(field, available_fields)
                        if alts:
                            suggestions[field] = alts

                for idx in range(n, step):
                    safe_set(ws, record_start + idx, cfg["spec_col"], None)

                if missing:
                    rating_issues.append({
                        "sheet": sheet_name,
                        "ref": item["ref"],
                        "part": part,
                        "cat": cat,
                        "sub": item["sub"],
                        "bom_row": item["bom_row"],
                        "missing_fields": missing,
                        "lookup_has_any": lookup_has_any,
                        "available_fields": sorted(available_fields),
                        "available_raw_fields": sorted(available_raw_fields),
                        "suggestions": suggestions
                    })
                continue

            # 일반 시트: 1개 정격값
            spec_one = ""
            for f in FIELD_ORDER_BY_CATEGORY.get(cat, []):
                if f in part_ratings and part_ratings[f]:
                    spec_one = part_ratings[f]
                    break
            safe_set(ws, record_start, cfg["spec_col"], spec_one if spec_one else None)

            if not spec_one:
                suggestions = {}
                if lookup_has_any:
                    for target in ["V_MAX", "I_MAX", "P_MAX", "V_RATED", "I_RATED"]:
                        alts = suggest_alternatives(target, available_fields)
                        if alts:
                            suggestions[target] = alts

                rating_issues.append({
                    "sheet": sheet_name,
                    "ref": item["ref"],
                    "part": part,
                    "cat": cat,
                    "sub": item["sub"],
                    "bom_row": item["bom_row"],
                    "missing_fields": ["(NO_MATCHED_FIELD)"],
                    "lookup_has_any": lookup_has_any,
                    "available_fields": sorted(available_fields),
                    "available_raw_fields": sorted(available_raw_fields),
                    "suggestions": suggestions
                })

    # ✅ [추가] 해당 시트에 부품이 0개면 템플릿 예시(첫 레코드) 값을 공란 처리
    for sheet_name in MANAGED_SHEETS:
        if len(out_groups.get(sheet_name, [])) == 0:
            ws = tpl_wb[sheet_name]
            start_row, step = layouts[sheet_name]
            cfg = SHEET_CFG[sheet_name]
            clear_first_record_values(ws, start_row, step, cfg)

    if unclassified_rows:
        write_unclassified_sheet(tpl_wb, ws_bom, unclassified_rows, UNCLASS_SHEET)

    tpl_wb.save(out_xlsx)
    write_issue_report(out_txt, duplicate_refs, rating_issues, routed_items)

    return {
        "out_xlsx": out_xlsx,
        "out_txt": out_txt,
        "ignored": dict(ignored),
        "written_counts": {k: len(v) for k, v in out_groups.items()},
    }


# =========================
# GUI (한 화면)
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BOM Parser")
        self.geometry("780x320")
        self.resizable(False, False)

        self.var_bom = tk.StringVar()
        self.var_template = tk.StringVar()
        self.var_lookup = tk.StringVar()
        self.var_outdir = tk.StringVar()
        self.var_status = tk.StringVar(value="Ready.")

        pad = {"padx": 10, "pady": 6}

        self._row_file("BOM (.xlsx)", self.var_bom, self.browse_bom, 0, **pad)
        self._row_file("Template (.xlsx)", self.var_template, self.browse_template, 1, **pad)
        self._row_file("LookupTable (.xlsx)", self.var_lookup, self.browse_lookup, 2, **pad)
        self._row_dir("Output Folder", self.var_outdir, self.browse_outdir, 3, **pad)

        frm_btn = tk.Frame(self)
        frm_btn.grid(row=4, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

        self.btn_run = tk.Button(frm_btn, text="Run", width=14, command=self.on_run)
        self.btn_run.pack(side="left")

        self.btn_clear = tk.Button(frm_btn, text="Clear", width=14, command=self.on_clear)
        self.btn_clear.pack(side="left", padx=10)

        self.btn_quit = tk.Button(frm_btn, text="Quit", width=14, command=self.destroy)
        self.btn_quit.pack(side="right")

        lbl_status = tk.Label(self, textvariable=self.var_status, anchor="w")
        lbl_status.grid(row=5, column=0, columnspan=3, sticky="ew", padx=10, pady=6)

        # ✅ EXE/스크립트 폴더에서 TEMPLATE/LOOKUPTABLE 자동 감지하여 프리필
        app_dir = get_app_dir()
        t, l = autodetect_default_files(app_dir)
        if t and not self.var_template.get():
            self.var_template.set(str(t))
        if l and not self.var_lookup.get():
            self.var_lookup.set(str(l))

    def _row_file(self, label, var, cmd, r, padx=10, pady=6):
        tk.Label(self, text=label, width=16, anchor="w").grid(row=r, column=0, padx=padx, pady=pady, sticky="w")
        tk.Entry(self, textvariable=var, width=75).grid(row=r, column=1, padx=padx, pady=pady, sticky="w")
        tk.Button(self, text="Browse...", width=12, command=cmd).grid(row=r, column=2, padx=padx, pady=pady)

    def _row_dir(self, label, var, cmd, r, padx=10, pady=6):
        tk.Label(self, text=label, width=16, anchor="w").grid(row=r, column=0, padx=padx, pady=pady, sticky="w")
        tk.Entry(self, textvariable=var, width=75).grid(row=r, column=1, padx=padx, pady=pady, sticky="w")
        tk.Button(self, text="Browse...", width=12, command=cmd).grid(row=r, column=2, padx=padx, pady=pady)

    def browse_bom(self):
        p = filedialog.askopenfilename(title="Select BOM", filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.var_bom.set(p)
            if not self.var_outdir.get():
                self.var_outdir.set(str(Path(p).parent))

    def browse_template(self):
        p = filedialog.askopenfilename(title="Select Template", filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.var_template.set(p)

    def browse_lookup(self):
        p = filedialog.askopenfilename(title="Select LookupTable", filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.var_lookup.set(p)

    def browse_outdir(self):
        p = filedialog.askdirectory(title="Select Output Folder")
        if p:
            self.var_outdir.set(p)

    def on_clear(self):
        self.var_bom.set("")
        self.var_template.set("")
        self.var_lookup.set("")
        self.var_outdir.set("")
        self.var_status.set("Ready.")

        # Clear 후에도 자동 감지 프리필 다시 채움
        app_dir = get_app_dir()
        t, l = autodetect_default_files(app_dir)
        if t:
            self.var_template.set(str(t))
        if l:
            self.var_lookup.set(str(l))

    def on_run(self):
        bom = Path(self.var_bom.get())
        tpl = Path(self.var_template.get())
        lk = Path(self.var_lookup.get())
        outdir = Path(self.var_outdir.get()) if self.var_outdir.get() else None

        if not bom.exists():
            messagebox.showerror("Error", "BOM 파일을 선택하세요.")
            return
        if not tpl.exists():
            messagebox.showerror("Error", "Template 파일을 선택하세요.")
            return
        if not lk.exists():
            messagebox.showerror("Error", "LookupTable 파일을 선택하세요.")
            return
        if outdir is None:
            outdir = bom.parent
        outdir.mkdir(parents=True, exist_ok=True)

        stem = bom.stem
        out_xlsx = outdir / f"{stem}_out.xlsx"
        out_txt = outdir / f"{stem}_issues.txt"

        try:
            self.var_status.set("Running...")
            self.update_idletasks()

            result = run_parser(bom, tpl, lk, out_xlsx, out_txt)

            msg = (
                f"완료!\n\n"
                f"- Output Excel: {result['out_xlsx']}\n"
                f"- Report TXT : {result['out_txt']}\n\n"
                f"- Written counts: {result['written_counts']}\n"
                f"- Ignored: {result['ignored']}\n"
            )
            self.var_status.set("Done.")
            messagebox.showinfo("Done", msg)

        except Exception as e:
            self.var_status.set("Error.")
            messagebox.showerror("Error", f"{e}")


# =========================
# CLI
# =========================
def parse_args():
    p = argparse.ArgumentParser(description="BOM Parser (Excel -> Template Derating)")
    p.add_argument("--bom", type=str, default="", help="Input BOM .xlsx path")
    p.add_argument("--template", type=str, default="", help="Template .xlsx path")
    p.add_argument("--lookup", type=str, default="", help="LookupTable .xlsx path")
    p.add_argument("--outdir", type=str, default="", help="Output directory")
    p.add_argument("--nogui", action="store_true", help="CLI only")
    return p.parse_args()


def main():
    args = parse_args()

    if args.nogui:
        if not args.bom or not args.template or not args.lookup:
            raise ValueError("CLI 모드에서는 --bom --template --lookup 을 모두 지정해야 합니다.")
        bom = Path(args.bom)
        tpl = Path(args.template)
        lk = Path(args.lookup)
        outdir = Path(args.outdir) if args.outdir else bom.parent
        outdir.mkdir(parents=True, exist_ok=True)

        out_xlsx = outdir / f"{bom.stem}_out.xlsx"
        out_txt = outdir / f"{bom.stem}_issues.txt"

        result = run_parser(bom, tpl, lk, out_xlsx, out_txt)

        print("DONE")
        print("Output:", result["out_xlsx"])
        print("Report:", result["out_txt"])
        return

    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
