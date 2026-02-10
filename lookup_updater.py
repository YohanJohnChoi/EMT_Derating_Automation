"""
LookupTable Updater - Windows GUI (Tkinter)

Features:
- Add/Update parts in LOOKUPTABLE.xlsx (TABLE sheet only)
- Download latest LOOKUPTABLE.xlsx from Google Drive (file link/ID)
"""

import re
import sys
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from io import FileIO
from urllib.request import urlopen, Request
from urllib.parse import urlparse, parse_qs

import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox


# =========================
# Helpers
# =========================
def normalize_text(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\u00A0", " ").replace("\u200B", "").replace("\ufeff", "")
    return s.strip()


def get_app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def autodetect_lookup(app_dir: Path) -> Path | None:
    xlsx = sorted(app_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for p in xlsx:
        if "LOOKUPTABLE" in p.name.upper():
            return p
    return None


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# =========================
# Lookup schema
# =========================
TABLE_SHEET = "TABLE"
ROUTING_SHEET = "ROUTING_RULES"
SCOPES = ["https://www.googleapis.com/auth/drive.file"]
CREDENTIALS_FILE = "credentials.json"
TOKEN_FILE = "token.json"
DRIVE_PENDING_FOLDER_ID = "1TAl_2hpp6HR08BvfIj81xSMlXrHBgwMM"


BASE_CAT_TO_SHEET = {
    "RESISTOR": "Resistor",
    "CAPACITOR": "Capacitor",
    "INDUCTOR": "Inductor",
    "FILTER": "Inductor",
    "DIODE": "Diode(ESD_Zener_Surge)",
    "TR": "FET&TR",
    "FET": "FET&TR",
    "POWER": "DCDC & LDO",
    "IC": "IC",
    "CONNECTOR": "Connector",
}


ALLOWED_CATS_BY_SHEET = {
    "Resistor": ["RESISTOR"],
    "Capacitor": ["CAPACITOR"],
    "Inductor": ["INDUCTOR", "FILTER"],
    "Diode(ESD_Zener_Surge)": ["DIODE"],
    "Diode(Schottky_switching)": ["DIODE"],
    "FET&TR": ["TR", "FET"],
    "DCDC & LDO": ["POWER"],
    "IC": ["IC"],
    "Connector": ["CONNECTOR"],
}


ALLOWED_FIELDS_BY_SHEET = {
    "Resistor": ["P_MAX", "V_MAX", "I_MAX", "V_RATED", "I_RATED"],
    "Capacitor": ["V_RATED", "V_MAX"],
    "Inductor": ["I_RATED", "I_MAX", "CURRENT", "DCR"],
    "Diode(ESD_Zener_Surge)": ["VRWM", "VBR_VPT", "V_MAX", "V_RATED"],
    "Diode(Schottky_switching)": ["V_MAX", "I_MAX"],
    "FET&TR": ["V_MAX", "I_MAX"],
    "DCDC & LDO": ["V_MAX", "I_MAX", "P_MAX"],
    "IC": ["V_MAX", "V_RATED", "I_MAX", "P_MAX"],
    "Connector": ["I_RATED", "I_MAX"],
}


REQUIRED_FIELDS_BY_SHEET = {
    "Diode(ESD_Zener_Surge)": ["VRWM", "VBR_VPT"],
    "Diode(Schottky_switching)": ["V_MAX", "I_MAX"],
    "FET&TR": ["V_MAX", "I_MAX"],
    "DCDC & LDO": ["V_MAX", "I_MAX", "P_MAX"],
}


@dataclass
class TableRowKey:
    category: str
    subcategory: str
    part_name: str
    rating_field: str


def load_lookup_schema(lookup_path: Path):
    wb = openpyxl.load_workbook(lookup_path, data_only=True)
    if TABLE_SHEET not in wb.sheetnames or ROUTING_SHEET not in wb.sheetnames:
        raise ValueError("LOOKUPTABLE.xlsx must contain TABLE and ROUTING_RULES sheets.")

    ws_table = wb[TABLE_SHEET]
    hdr = {normalize_text(ws_table.cell(1, c).value): c for c in range(1, ws_table.max_column + 1)}
    required_cols = ["Category", "Subcategory", "Part_Name", "Rating_Field", "Rating_Value", "Rating_Unit"]
    for req in required_cols:
        if req not in hdr:
            raise ValueError(f"TABLE sheet missing column: {req}")

    ws_rules = wb[ROUTING_SHEET]
    rules = []
    for r in range(2, ws_rules.max_row + 1):
        cat = normalize_text(ws_rules.cell(r, 1).value)
        sub = normalize_text(ws_rules.cell(r, 2).value)
        sheet = normalize_text(ws_rules.cell(r, 3).value)
        if not cat and not sub and not sheet:
            continue
        rules.append((cat.upper(), sub, sheet))

    return wb, ws_table, hdr, rules


def build_subcategory_map(wb) -> dict:
    ws = wb[TABLE_SHEET]
    hdr = {normalize_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    sub_map = {}
    for r in range(2, ws.max_row + 1):
        cat = normalize_text(ws.cell(r, hdr["Category"]).value).upper()
        sub = normalize_text(ws.cell(r, hdr["Subcategory"]).value)
        if not cat:
            continue
        sub_map.setdefault(cat, set()).add(sub)
    return {k: sorted(v) for k, v in sub_map.items()}


def build_row_index(ws, hdr) -> dict:
    index = {}
    for r in range(2, ws.max_row + 1):
        cat = normalize_text(ws.cell(r, hdr["Category"]).value).upper()
        sub = normalize_text(ws.cell(r, hdr["Subcategory"]).value)
        part = normalize_text(ws.cell(r, hdr["Part_Name"]).value)
        field = normalize_text(ws.cell(r, hdr["Rating_Field"]).value).upper()
        if not cat or not part or not field:
            continue
        index[TableRowKey(cat, sub, part, field)] = r
    return index


# =========================
# Google Drive Download
# =========================
def extract_drive_file_id(link: str) -> str | None:
    link = link.strip()
    if not link:
        return None

    # Direct file ID
    if re.fullmatch(r"[a-zA-Z0-9_-]{10,}", link):
        return link

    if "drive.google.com" not in link:
        return None

    if "/file/d/" in link:
        m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", link)
        return m.group(1) if m else None

    parsed = urlparse(link)
    q = parse_qs(parsed.query)
    if "id" in q and q["id"]:
        return q["id"][0]

    return None


def is_drive_folder_link(link: str) -> bool:
    return "drive.google.com" in link and "folders" in link


def extract_drive_folder_id(link: str) -> str | None:
    if "drive.google.com" not in link:
        return None
    m = re.search(r"/folders/([a-zA-Z0-9_-]+)", link)
    return m.group(1) if m else None


def download_drive_file(file_id: str, dest_path: Path):
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req) as resp:
        data = resp.read()

    # Handle confirm token if Google warns about large files
    if b"confirm=" in data and b"uc?export=download" in data:
        m = re.search(rb"confirm=([0-9A-Za-z_]+)", data)
        if m:
            token = m.group(1).decode("utf-8")
            url2 = f"https://drive.google.com/uc?export=download&confirm={token}&id={file_id}"
            req2 = Request(url2, headers={"User-Agent": "Mozilla/5.0"})
            with urlopen(req2) as resp2:
                data = resp2.read()

    dest_path.write_bytes(data)


def get_drive_service(app_dir: Path):
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request as GoogleRequest
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except Exception as e:
        raise RuntimeError(
            "Google Drive dependencies missing. Install: "
            "google-api-python-client google-auth-httplib2 google-auth-oauthlib"
        ) from e

    cred_path = app_dir / CREDENTIALS_FILE
    if not cred_path.exists():
        raise FileNotFoundError(f"Missing {CREDENTIALS_FILE} in: {app_dir}")

    token_path = app_dir / TOKEN_FILE
    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(GoogleRequest())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(cred_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds)


def get_latest_file_in_folder(service, folder_id: str) -> dict | None:
    q = f"'{folder_id}' in parents and trashed = false"
    resp = service.files().list(
        q=q,
        orderBy="modifiedTime desc",
        fields="files(id,name,modifiedTime,mimeType)",
        pageSize=50,
    ).execute()
    files = resp.get("files", [])
    if not files:
        return None

    for f in files:
        if f.get("name", "").lower().endswith(".xlsx"):
            return f
    return files[0]


def download_drive_file_via_api(service, file_id: str, dest_path: Path):
    from googleapiclient.http import MediaIoBaseDownload

    req = service.files().get_media(fileId=file_id)
    with FileIO(dest_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()


def upload_drive_file(service, file_path: Path, folder_id: str, mime_type: str | None = None):
    from googleapiclient.http import MediaFileUpload

    body = {"name": file_path.name, "parents": [folder_id]}
    media = MediaFileUpload(str(file_path), mimetype=mime_type, resumable=False)
    req = service.files().create(body=body, media_body=media, fields="id")
    return req.execute()


def format_upload_report(records: list, lookup_path: Path) -> str:
    lines = []
    lines.append("=== LookupTable Upload Report ===")
    lines.append(f"- Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- Source: {lookup_path}")
    lines.append("")

    added = [r for r in records if r["action"] == "added"]
    updated = [r for r in records if r["action"] == "updated"]

    lines.append("[1] Added Records")
    if not added:
        lines.append("  - None")
    else:
        for r in added:
            sub = r["subcategory"] if r["subcategory"] else "(blank)"
            lines.append(
                f"  - Part={r['part']}, Category={r['category']}, Subcategory={sub}, "
                f"Field={r['field']}, Value={r['value']}, Unit={r['unit']}"
            )
    lines.append("")

    lines.append("[2] Updated Records")
    if not updated:
        lines.append("  - None")
    else:
        for r in updated:
            sub = r["subcategory"] if r["subcategory"] else "(blank)"
            lines.append(
                f"  - Part={r['part']}, Category={r['category']}, Subcategory={sub}, "
                f"Field={r['field']}, Value={r['value']}, Unit={r['unit']}"
            )
    lines.append("")

    return "\n".join(lines)


# =========================
# GUI
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("LookupTable Updater")
        self.geometry("940x640")
        self.resizable(False, False)

        self.var_lookup = tk.StringVar()
        self.var_sheet = tk.StringVar()
        self.var_category = tk.StringVar()
        self.var_subcategory = tk.StringVar()
        self.var_part = tk.StringVar()
        self.var_drive_link = tk.StringVar()
        self.var_status = tk.StringVar(value="Ready.")

        self.rating_entries = {}
        self.allowed_fields = []
        self.subcat_map = {}
        self.change_records = []

        pad = {"padx": 10, "pady": 6}

        self._row_file("LookupTable (.xlsx)", self.var_lookup, self.browse_lookup, 0, **pad)
        self._row_text("Drive file link/ID", self.var_drive_link, 1, **pad)

        frm_down = tk.Frame(self)
        frm_down.grid(row=2, column=0, columnspan=3, sticky="ew", padx=10, pady=6)
        tk.Button(frm_down, text="Download Latest", width=18, command=self.on_download).pack(side="left")
        tk.Button(frm_down, text="Upload to Drive", width=18, command=self.on_upload).pack(side="left", padx=8)
        tk.Button(frm_down, text="Reload Lookup", width=18, command=self.on_reload).pack(side="left", padx=8)

        tk.Label(self, text="Part Add / Update", anchor="w").grid(
            row=3, column=0, columnspan=3, sticky="w", padx=10, pady=10
        )

        self._row_dropdown("Output Sheet", self.var_sheet, 4, self.on_sheet_changed, **pad)
        self._row_dropdown("Category", self.var_category, 5, self.on_category_changed, **pad)
        self._row_dropdown("Subcategory", self.var_subcategory, 6, None, **pad)
        self._row_text("Part Name", self.var_part, 7, **pad)

        self.frm_ratings = tk.Frame(self)
        self.frm_ratings.grid(row=8, column=0, columnspan=3, sticky="ew", padx=10, pady=6)

        frm_btn = tk.Frame(self)
        frm_btn.grid(row=9, column=0, columnspan=3, sticky="ew", padx=10, pady=10)
        self.btn_add = tk.Button(frm_btn, text="Add / Update", width=16, command=self.on_add)
        self.btn_add.pack(side="left")
        tk.Button(frm_btn, text="Clear", width=14, command=self.on_clear).pack(side="left", padx=8)
        tk.Button(frm_btn, text="Quit", width=14, command=self.destroy).pack(side="right")

        lbl_status = tk.Label(self, textvariable=self.var_status, anchor="w")
        lbl_status.grid(row=10, column=0, columnspan=3, sticky="ew", padx=10, pady=6)

        app_dir = get_app_dir()
        lk = autodetect_lookup(app_dir)
        if lk:
            self.var_lookup.set(str(lk))

        # Default drive link (editable)
        self.var_drive_link.set("")

        self.on_reload()

    def _row_file(self, label, var, cmd, r, padx=10, pady=6):
        tk.Label(self, text=label, width=18, anchor="w").grid(row=r, column=0, padx=padx, pady=pady, sticky="w")
        tk.Entry(self, textvariable=var, width=88).grid(row=r, column=1, padx=padx, pady=pady, sticky="w")
        tk.Button(self, text="Browse...", width=12, command=cmd).grid(row=r, column=2, padx=padx, pady=pady)

    def _row_text(self, label, var, r, padx=10, pady=6):
        tk.Label(self, text=label, width=18, anchor="w").grid(row=r, column=0, padx=padx, pady=pady, sticky="w")
        tk.Entry(self, textvariable=var, width=88).grid(row=r, column=1, padx=padx, pady=pady, sticky="w")

    def _row_dropdown(self, label, var, r, callback, padx=10, pady=6):
        tk.Label(self, text=label, width=18, anchor="w").grid(row=r, column=0, padx=padx, pady=pady, sticky="w")
        opt = tk.OptionMenu(self, var, "")
        opt.config(width=60)
        opt.grid(row=r, column=1, padx=padx, pady=pady, sticky="w")
        if callback:
            var.trace_add("write", lambda *args: callback())
        setattr(self, f"_opt_{label.replace(' ', '_').lower()}", opt)

    def browse_lookup(self):
        p = filedialog.askopenfilename(title="Select LookupTable", filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.var_lookup.set(p)
            self.on_reload()

    def on_reload(self):
        lk = Path(self.var_lookup.get())
        if not lk.exists():
            self.var_status.set("LookupTable not found.")
            return

        try:
            wb, _, _, _ = load_lookup_schema(lk)
            self.subcat_map = build_subcategory_map(wb)

            self._set_sheet_options(sorted(ALLOWED_CATS_BY_SHEET.keys()))
            self.var_status.set("LookupTable loaded.")

        except Exception as e:
            self.var_status.set("Load failed.")
            messagebox.showerror("Error", f"{e}")

    def _set_sheet_options(self, sheets):
        opt = getattr(self, "_opt_output_sheet")
        menu = opt["menu"]
        menu.delete(0, "end")
        for s in sheets:
            menu.add_command(label=s, command=lambda v=s: self.var_sheet.set(v))
        if sheets:
            self.var_sheet.set(sheets[0])

    def _set_category_options(self, cats):
        opt = getattr(self, "_opt_category")
        menu = opt["menu"]
        menu.delete(0, "end")
        for c in cats:
            menu.add_command(label=c, command=lambda v=c: self.var_category.set(v))
        if cats:
            self.var_category.set(cats[0])
        else:
            self.var_category.set("")

    def _set_subcategory_options(self, subs):
        opt = getattr(self, "_opt_subcategory")
        menu = opt["menu"]
        menu.delete(0, "end")
        for s in subs:
            menu.add_command(label=s, command=lambda v=s: self.var_subcategory.set(v))
        if subs:
            self.var_subcategory.set(subs[0])
        else:
            self.var_subcategory.set("")

    def on_sheet_changed(self):
        sheet = self.var_sheet.get()
        cats = ALLOWED_CATS_BY_SHEET.get(sheet, [])
        self._set_category_options(cats)
        self._build_rating_fields(sheet)

    def on_category_changed(self):
        cat = self.var_category.get()
        subs = self.subcat_map.get(cat, [])
        if "" not in subs:
            subs = [""] + subs
        self._set_subcategory_options(subs)

    def _build_rating_fields(self, sheet):
        for widget in self.frm_ratings.winfo_children():
            widget.destroy()
        self.rating_entries.clear()

        fields = ALLOWED_FIELDS_BY_SHEET.get(sheet, [])
        self.allowed_fields = fields

        tk.Label(self.frm_ratings, text="Rating Field", width=20, anchor="w").grid(row=0, column=0, padx=6)
        tk.Label(self.frm_ratings, text="Value", width=18, anchor="w").grid(row=0, column=1, padx=6)
        tk.Label(self.frm_ratings, text="Unit", width=12, anchor="w").grid(row=0, column=2, padx=6)

        for i, f in enumerate(fields, start=1):
            tk.Label(self.frm_ratings, text=f, width=20, anchor="w").grid(row=i, column=0, padx=6, pady=3)
            v = tk.Entry(self.frm_ratings, width=20)
            u = tk.Entry(self.frm_ratings, width=12)
            v.grid(row=i, column=1, padx=6, pady=3, sticky="w")
            u.grid(row=i, column=2, padx=6, pady=3, sticky="w")
            self.rating_entries[f] = (v, u)

    def on_clear(self):
        self.var_sheet.set("")
        self.var_category.set("")
        self.var_subcategory.set("")
        self.var_part.set("")
        self._build_rating_fields("")
        self.var_status.set("Cleared.")
        self.on_reload()

    def on_add(self):
        lk = Path(self.var_lookup.get())
        if not lk.exists():
            messagebox.showerror("Error", "LookupTable file not found.")
            return

        sheet = self.var_sheet.get()
        cat = self.var_category.get()
        sub = self.var_subcategory.get()
        part = normalize_text(self.var_part.get())

        if not sheet or not cat or not part:
            messagebox.showerror("Error", "Output Sheet, Category, and Part Name are required.")
            return

        allowed_fields = ALLOWED_FIELDS_BY_SHEET.get(sheet, [])
        required_fields = REQUIRED_FIELDS_BY_SHEET.get(sheet, [])

        filled = {}
        for f, (v_ent, u_ent) in self.rating_entries.items():
            val = normalize_text(v_ent.get())
            unit = normalize_text(u_ent.get())
            if val:
                filled[f] = (val, unit)

        if required_fields:
            missing = [f for f in required_fields if f not in filled]
            if missing:
                messagebox.showerror("Error", f"Missing required fields: {', '.join(missing)}")
                return
        else:
            if not filled:
                messagebox.showerror("Error", "At least one rating field is required.")
                return

        try:
            wb, ws_table, hdr, _ = load_lookup_schema(lk)
            row_index = build_row_index(ws_table, hdr)

            updated = 0
            appended = 0
            for field, (val, unit) in filled.items():
                key = TableRowKey(cat, sub, part, field)
                if key in row_index:
                    r = row_index[key]
                    ws_table.cell(r, hdr["Rating_Value"]).value = val
                    ws_table.cell(r, hdr["Rating_Unit"]).value = unit
                    self.change_records.append({
                        "action": "updated",
                        "category": cat,
                        "subcategory": sub,
                        "part": part,
                        "field": field,
                        "value": val,
                        "unit": unit,
                    })
                    updated += 1
                else:
                    r = ws_table.max_row + 1
                    ws_table.cell(r, hdr["Category"]).value = cat
                    ws_table.cell(r, hdr["Subcategory"]).value = sub
                    ws_table.cell(r, hdr["Part_Name"]).value = part
                    ws_table.cell(r, hdr["Rating_Field"]).value = field
                    ws_table.cell(r, hdr["Rating_Value"]).value = val
                    ws_table.cell(r, hdr["Rating_Unit"]).value = unit
                    self.change_records.append({
                        "action": "added",
                        "category": cat,
                        "subcategory": sub,
                        "part": part,
                        "field": field,
                        "value": val,
                        "unit": unit,
                    })
                    appended += 1

            backup = lk.with_name(f"{lk.stem}_backup_{now_stamp()}{lk.suffix}")
            shutil.copy2(lk, backup)
            wb.save(lk)

            self.var_status.set("Updated.")
            messagebox.showinfo(
                "Done",
                f"Saved to: {lk}\nBackup: {backup}\nUpdated: {updated}, Added: {appended}",
            )

        except Exception as e:
            self.var_status.set("Error.")
            messagebox.showerror("Error", f"{e}")

    def on_download(self):
        link = self.var_drive_link.get().strip()
        if not link:
            messagebox.showerror("Error", "Enter a Google Drive file link or file ID.")
            return

        lk = Path(self.var_lookup.get())
        if not lk.exists():
            messagebox.showerror("Error", "LookupTable path is invalid.")
            return

        try:
            tmp = lk.with_name(f"{lk.stem}_download_{now_stamp()}{lk.suffix}")
            self.var_status.set("Downloading...")
            self.update_idletasks()

            if is_drive_folder_link(link):
                folder_id = extract_drive_folder_id(link)
                if not folder_id:
                    raise ValueError("Unable to extract folder ID from link.")
                service = get_drive_service(get_app_dir())
                latest = get_latest_file_in_folder(service, folder_id)
                if not latest:
                    raise ValueError("No files found in folder.")
                download_drive_file_via_api(service, latest["id"], tmp)
            else:
                file_id = extract_drive_file_id(link)
                if not file_id:
                    raise ValueError("Unable to extract file ID from link.")
                download_drive_file(file_id, tmp)

            backup = lk.with_name(f"{lk.stem}_backup_{now_stamp()}{lk.suffix}")
            shutil.copy2(lk, backup)
            shutil.move(tmp, lk)

            self.var_status.set("Downloaded.")
            messagebox.showinfo("Done", f"Updated: {lk}\nBackup: {backup}")
            self.on_reload()

        except Exception as e:
            self.var_status.set("Download failed.")
            messagebox.showerror("Error", f"{e}")

    def on_upload(self):
        lk = Path(self.var_lookup.get())
        if not lk.exists():
            messagebox.showerror("Error", "LookupTable path is invalid.")
            return

        if not self.change_records:
            if not messagebox.askyesno(
                "Upload",
                "No add/update records in this session.\nUpload anyway with empty report?",
            ):
                return

        try:
            app_dir = get_app_dir()
            report_txt = app_dir / f"lookup_upload_{now_stamp()}.txt"
            report_txt.write_text(format_upload_report(self.change_records, lk), encoding="utf-8")

            self.var_status.set("Uploading...")
            self.update_idletasks()

            service = get_drive_service(app_dir)
            upload_drive_file(
                service,
                lk,
                DRIVE_PENDING_FOLDER_ID,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            upload_drive_file(
                service,
                report_txt,
                DRIVE_PENDING_FOLDER_ID,
                mime_type="text/plain",
            )

            self.var_status.set("Uploaded.")
            messagebox.showinfo("Done", "LookupTable and report uploaded to Drive.")

        except Exception as e:
            self.var_status.set("Upload failed.")
            messagebox.showerror("Error", f"{e}")


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
